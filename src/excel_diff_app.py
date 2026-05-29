"""Streamlit app for comparing two Excel workbooks and creating replacement files."""

from __future__ import annotations

import copy
import difflib
import html
import io
from dataclasses import dataclass
from typing import Any, BinaryIO, TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook


@dataclass(frozen=True)
class SheetDiff:
    """Describes a structural difference between two sheets."""

    sheet: str
    diff_type: str
    left: str
    right: str


@dataclass(frozen=True)
class CellDiff:
    """Describes a cell value difference."""

    sheet: str
    cell: str
    diff_type: str
    left: Any
    right: Any
    char_diff_html: str


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def build_char_diff_html(left: Any, right: Any) -> str:
    """Build a compact HTML character-level diff for two values."""
    left_text = _stringify(left)
    right_text = _stringify(right)
    if left_text == right_text:
        return ""

    matcher = difflib.SequenceMatcher(None, left_text, right_text)
    chunks: list[str] = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        old = left_text[i1:i2]
        new = right_text[j1:j2]
        if tag == "equal":
            chunks.append(html.escape(old))
        elif tag == "delete":
            chunks.append(f"<del>{html.escape(old)}</del>")
        elif tag == "insert":
            chunks.append(f"<ins>{html.escape(new)}</ins>")
        elif tag == "replace":
            chunks.append(f"<del>{html.escape(old)}</del><ins>{html.escape(new)}</ins>")
    return "".join(chunks)


def get_column_letter(index: int) -> str:
    """Convert a 1-based column index to an Excel column letter."""
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def load_excel(source: BinaryIO | bytes) -> "Workbook":
    """Load an Excel workbook from an uploaded file or bytes."""
    from openpyxl import load_workbook

    if isinstance(source, bytes):
        source = io.BytesIO(source)
    return load_workbook(source)


def compare_workbooks(left_wb: "Workbook", right_wb: "Workbook") -> tuple[list[SheetDiff], list[CellDiff]]:
    """Compare workbook structure and cell values."""
    sheet_diffs: list[SheetDiff] = []
    cell_diffs: list[CellDiff] = []

    left_sheets = set(left_wb.sheetnames)
    right_sheets = set(right_wb.sheetnames)

    for sheet in sorted(left_sheets - right_sheets):
        sheet_diffs.append(SheetDiff(sheet, "工作表仅在左侧存在", "存在", "缺失"))
    for sheet in sorted(right_sheets - left_sheets):
        sheet_diffs.append(SheetDiff(sheet, "工作表仅在右侧存在", "缺失", "存在"))

    for sheet in sorted(left_sheets & right_sheets):
        left_ws = left_wb[sheet]
        right_ws = right_wb[sheet]
        left_size = f"{left_ws.max_row} 行 × {left_ws.max_column} 列"
        right_size = f"{right_ws.max_row} 行 × {right_ws.max_column} 列"
        if left_size != right_size:
            sheet_diffs.append(SheetDiff(sheet, "工作表尺寸不同", left_size, right_size))

        max_row = max(left_ws.max_row, right_ws.max_row)
        max_col = max(left_ws.max_column, right_ws.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                left_value = left_ws.cell(row=row, column=col).value
                right_value = right_ws.cell(row=row, column=col).value
                if left_value == right_value:
                    continue

                cell = f"{get_column_letter(col)}{row}"
                if left_value is None and right_value is not None:
                    diff_type = "右侧新增数据"
                elif left_value is not None and right_value is None:
                    diff_type = "右侧删除数据"
                elif type(left_value) is not type(right_value):
                    diff_type = "数据类型不同"
                else:
                    diff_type = "值不同"

                cell_diffs.append(
                    CellDiff(
                        sheet=sheet,
                        cell=cell,
                        diff_type=diff_type,
                        left=left_value,
                        right=right_value,
                        char_diff_html=build_char_diff_html(left_value, right_value),
                    )
                )

    return sheet_diffs, cell_diffs


def _copy_sheet_values(source_ws, target_wb: Workbook) -> None:
    target_ws = target_wb.create_sheet(source_ws.title)
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(row=source_cell.row, column=source_cell.column)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy.copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.comment:
                target_cell.comment = copy.copy(source_cell.comment)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy.copy(source_cell.hyperlink)

    for column, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[column].width = dimension.width
    for row_index, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_index].height = dimension.height


def create_replacement_workbook_bytes(source_wb: "Workbook", replacement_wb: "Workbook") -> bytes:
    """Create a new workbook that replaces source workbook sheets with replacement workbook content."""
    from openpyxl.workbook import Workbook

    del source_wb  # Replacement output is intentionally based on the selected replacement file.
    output_wb = Workbook()
    default_sheet = output_wb.active
    output_wb.remove(default_sheet)

    for source_ws in replacement_wb.worksheets:
        _copy_sheet_values(source_ws, output_wb)

    output = io.BytesIO()
    output_wb.save(output)
    return output.getvalue()


def sheet_diffs_to_frame(sheet_diffs: list[SheetDiff]):
    import pandas as pd

    return pd.DataFrame([diff.__dict__ for diff in sheet_diffs])


def cell_diffs_to_frame(cell_diffs: list[CellDiff]):
    import pandas as pd

    return pd.DataFrame([diff.__dict__ for diff in cell_diffs])


def render_app() -> None:
    import streamlit as st

    st.set_page_config(page_title="Excel 文档对比与替换工具", layout="wide")
    st.title("Excel 文档对比与替换工具")
    st.caption("上传两个 Excel 文件，查看结构、数据和字符级差异，并生成一键替换后的文件。")

    left_file = st.file_uploader("左侧/原始 Excel", type=["xlsx", "xlsm", "xltx", "xltm"], key="left")
    right_file = st.file_uploader("右侧/目标 Excel", type=["xlsx", "xlsm", "xltx", "xltm"], key="right")

    if not left_file or not right_file:
        st.info("请上传两个 Excel 文件开始对比。")
        return

    left_bytes = left_file.getvalue()
    right_bytes = right_file.getvalue()
    left_wb = load_excel(left_bytes)
    right_wb = load_excel(right_bytes)
    sheet_diffs, cell_diffs = compare_workbooks(left_wb, right_wb)

    st.subheader("结构差异")
    if sheet_diffs:
        st.dataframe(sheet_diffs_to_frame(sheet_diffs), use_container_width=True)
    else:
        st.success("未发现工作表结构差异。")

    st.subheader("数据差异")
    if cell_diffs:
        cell_df = cell_diffs_to_frame(cell_diffs)
        st.dataframe(cell_df.drop(columns=["char_diff_html"]), use_container_width=True)
        st.download_button(
            "下载差异明细 CSV",
            data=cell_df.to_csv(index=False).encode("utf-8-sig"),
            file_name="excel-diff-details.csv",
            mime="text/csv",
        )

        st.subheader("字符级差异")
        selected = st.selectbox(
            "选择差异单元格",
            options=range(len(cell_diffs)),
            format_func=lambda index: f"{cell_diffs[index].sheet}!{cell_diffs[index].cell}",
        )
        st.markdown(
            "<style>del{background:#ffd6d6;} ins{background:#d6ffd6;text-decoration:none;}</style>",
            unsafe_allow_html=True,
        )
        st.markdown(cell_diffs[selected].char_diff_html or "无字符级差异", unsafe_allow_html=True)
    else:
        st.success("未发现单元格数据差异。")

    st.subheader("一键修改与替换")
    left_to_right = create_replacement_workbook_bytes(left_wb, right_wb)
    right_to_left = create_replacement_workbook_bytes(right_wb, left_wb)
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "下载：左侧替换为右侧内容",
            data=left_to_right,
            file_name="left-replaced-by-right.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col2:
        st.download_button(
            "下载：右侧替换为左侧内容",
            data=right_to_left,
            file_name="right-replaced-by-left.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    render_app()
